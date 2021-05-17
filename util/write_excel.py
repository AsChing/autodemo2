
# import os, sys
#
# sys.path.append(os.path.dirname(os.path.dirname(__file__)))
import shutil
from autodemo.config.public_data import *
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
import configparser as cparser
from log import my_log


class WriteExcel(object):

    def __init__(self, fileName, SheetName="Sheet1"):
        self.filename = fileName
        if not os.path.exists(self.filename):
            # 文件不存在，则拷贝文件至指定目录下
            shutil.copyfile(FILE_PATH, REPORT_FILE_PATH)
        self.wb = load_workbook(self.filename)
        self.ws = self.wb[SheetName]

        # --------- 读取config.ini配置文件 ---------------
        cf = cparser.ConfigParser()
        cf.read("../config/config.ini", encoding='UTF-8')
        self.name = cf.get("tester", "name")
        self.logger = my_log.Logger('../log/log/all.log', level='debug').logger
        self.error_log = my_log.Logger('../log/log/error.log', level='error').logger

    def write_data(self, row_n, result, response):
        try:
            font_GREEN = Font(name='宋体', color="ff10af06", bold=True, size=14)
            font_RED = Font(name='宋体', color="fff10f0f", bold=True, size=14)
            font_Name = Font(name='宋体', color="ff5d0679", bold=True, size=12)
            align = Alignment(horizontal='center', vertical='center')
            # 获数所在行数
            L_n = "L" + str(row_n)
            N_n = "N" + str(row_n)
            # 写入测试结果
            if result == "PASS":
                self.ws.cell(row_n, 12, result)
                self.ws[L_n].font = font_GREEN
            if result == "FAIL":
                self.ws.cell(row_n, 12, result)
                self.ws[L_n].font = font_RED
            # 写入接口返回code
            self.ws.cell(row_n, 13, response)
            # 写入测试人员名称
            self.ws.cell(row_n, 14, self.name)
            self.ws[L_n].alignment = align
            self.ws[N_n].font = font_Name
            self.ws[N_n].alignment = align
            self.wb.save(self.filename)
            self.logger.info("----------------write_data成功--------------")
        except Exception as e:
            self.logger.warning("-----------write_data失败------------")
            self.error_log.error("写入测试结果数据报错了---", e)

    #     某一格写入数据
    def write_somewhere(self, row, result, res):
        try:
            self.ws.cell(row, 6, result)
            self.ws.cell(row, 5, res)
            self.wb.save(self.filename)
            self.logger.info("----------------write_{0}行流程接口结果成功--------------".format(row))
        except Exception as e:
            self.logger.warning("-----------write_{0}行流程接口结果失败------------".format(row))
            self.error_log.error("写入测试结果数据报错了---", e)


if __name__ == '__main__':
    we = WriteExcel(FILE_PATH)
    # we.write_data(2, "FAIL", "{'code':400}")
    # we.write_data(3, "PASS", "{'code':200}")
    we.write_somewhere(3, 3, "测试数据")
